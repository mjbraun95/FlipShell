from openpyxl import Workbook
import datetime

def yesOrNo(msg):
    while True:
        userIn = ""
        userIn = input(msg + " [Y/n]: ").lower()
        if userIn == "y":
            return True
        elif userIn == "n":
            return False

class Listing:
    def __init__(self):
        self.platform
        self.title
        self.description
        self.price
        self.availableQuantity
        self.soldQuantity
        self.lastDateTimeChecked

    def respondToOffers(self):
        #Open url to messages then confirm
        return

    def markAsSold(self):
        return

def openURL(url):
    return

class Item:
    def __init__(self):
        self.name = input("Item name: ")
        self.prevStatus = self.status
        self.updateStatus()

        self.updateInfo()
        self.update()

    def updateStatus(self):
        options = {
            "u" : "uncopped",
            "c" : "copped",
            "i" : "in hand",
            "l" : "listed",
            "p" : "paid for",
            "s" : "shipped",
            "d" : "delivered",
            "c" : "completed",
        }
        statusSelect = input("Input one of the following letters \n" + 
        "    U = Uncopped,\n" + 
        "    C = Copped, not arrived\n" + 
        "    I = In hand, not listed\n" + 
        "    L = Listed, not sold\n" + 
        "    P = Paid + sold, not shipped\n" + 
        "    S = Shipped, not delivered\n" + 
        "    D = Delivered\n" + 
        "    C = Completed\n" +
        "\nItem status: ").lower()
        print("options[statusSelect]: {}".format(options[statusSelect]))


        self.status = options[statusSelect]

    def updateInfo(self):
        if self.status == "u":

        elif self.status == "c":
        elif self.status == "i":
        elif self.status == "l":
        elif self.status == "p":
        elif self.status == "s":
        elif self.status == "d":
        elif self.status == "c":
        
        self.copSite = input("Cop Website: ")
        self.copPrice = float( input("Cop Price: ") )
        self.copped = yesOrNo("Copped?")
        self.trackingURL = -1
        self.inHandDate = -1
        self.inHand = -1
        self.listed = False
        self.listings = -1

    def update(self):
        self.addedToBots = yesOrNo("Added To Bots?")
        self.copped = yesOrNo("Copped?")
        self.inHand = yesOrNo("In hand?")
        self.sites = []
        for site in {"eBay", "Facebook", "Kijiji", "GOAT", "StockX"}:
            siteIn = yesOrNo("List on {}?".format(site))
            if siteIn:
                self.sites.append(siteIn)
        self.sites = yesOrNo("Listed?")
        self.listed = yesOrNo("Listed?")
        self.paid = yesOrNo("Sold + Paid?")
        self.shipped = yesOrNo("Shipped?")
        self.completed = yesOrNo("Completed?")
        if self.addedToBots == False:
            self.taskAtHand = "Add drop info to bots"
            return
        if self.copped == False:

            self.copYear = input("Cop Year (Enter for 2020) [yyyy]: ")
            if self.copYear == "":
                self.copYear = 2020
            else:
                self.copYear = int(self.copYear)
            
            self.copMonth = int(input("Cop Month Number: "))
            self.copDay = int(input("Cop Day Number: "))
            self.copHour = int(input("Cop Hour Number (24-hour MST): "))
            self.copMinute = int(input("Cop Minute Number: "))
            self.copDateTime = datetime.datetime(self.copYear, self.copMonth, self.copDay, self.copHour, self.copMinute)
            if self.copDateTime < datetime.datetime.now():
                print("Too late! Current date = {}".format(datetime.datetime.now()))
            self.taskAtHand = "Wait for drop"
            return
        else:
            if self.trackingURL == -1:
                self.trackingURL = input("Tracking URL: ")
            else:
                track = yesOrNo("Open tracking URL?")
                if track == True:
                    openURL()
            self.inHandDate = input("Latest expected delivery date: ")

        if self.inHand == True and self.listed == False:
            self.listAds()
        elif self.inHand == False and self.listed == False:
            listAds = yesOrNo("List ads?")
            if listAds == True:
                self.listAds()
        elif self.listed == True:
            checkOnAds = yesOrNo("Check on ads?")
            if checkOnAds == True:
                self.checkOnAds()

    def processAddToBot(self):
        return

    def processCop(self):
        return

    def processInHand(self):
        return

    def processCop(self):
        return

    def listAds(self):
        onEbay = yesOrNo("On eBay?")
        onKijiji = yesOrNo("On Kijiji?")
        onMarketplace = yesOrNo("On Marketplace?")
        onStockX = yesOrNo("On StockX?")
        #Take user through data fill
        #Add to self.listings
        return

    def checkOnAds(self):
        #Ask to remove ads
        #Take user through data fill
        #Add to self.listings
        return

class Sheet:
    def __init__(self, filename = None):
        if filename is None:
            filename = input("Input sheet name (no extension): ")
        self.filename = filename
        self.items = []
        self.listings = []

    def addItem(self):
        newItem = Item()
        self.items.append(newItem)
        print("{} items now".format(len(self.items)))

    def addListing(self):
        newListing = Listing()
        self.listings.append(newListing)
        print("{} listings now".format(len(self.listings)))

if __name__ == "__main__":
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['A1'] = 42

    # Rows can also be appended
    ws.append([1, 2, 3])

    # Python types will automatically be converted
    
    ws['A2'] = datetime.datetime.now()

    # Save the file
    wb.save("auto.xlsx")